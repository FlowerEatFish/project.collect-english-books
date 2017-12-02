"""Read and parser content of Excel file."""
from openpyxl import load_workbook

class CollectAllIsbn(object):
    """path: file path(string), return: isbn_list(list[int])"""
    def __init__(self, path):
        self.path = str(path)
        self.max_collection = 100 # int, 最大收集數

        if self.check_file(self.path):
            self.file = load_workbook(filename=self.path)
            self.data = self.collect_isbn()
        else:
            print('The path is incorrect.')
            self.data = 'No data due to file not found.'

    @classmethod
    def check_file(cls, path):
        """path: string, return: bool"""
        try:
            load_workbook(filename=path)
            return True
        except:
            return False

    def check_first_column(self):
        """No parameter is required, return: string"""
        sheet = self.file[self.file.get_sheet_names()[0]]
        for i in range(50):
            check_block = sheet['%s1' % chr(ord('A')+i)]
            if check_block.value == 'ISBN' or check_block.value == 'isbn':
                return '%s' % chr(ord('A')+i)

    def collect_isbn(self):
        """No parameter is required, return: list[int]"""
        column = self.check_first_column()
        sheet = self.file[self.file.get_sheet_names()[0]]

        isbn_list = []
        location_null_data = 0
        for i in range(self.max_collection):
            if location_null_data > 2:
                break
            isbn = sheet['%s%d' % (column, (i+2))].value
            if isbn is None:
                location_null_data += 1
                isbn_list.append('None')
            else:
                location_null_data = 0
                isbn_list.append(isbn)
        return isbn_list

# Demo
if __name__ == '__main__':
    PATH = 'test.xlsx'
    DATA = CollectAllIsbn(PATH)
    print(DATA.data)
